#!/usr/bin/env python3
"""
Interactively create a Matrixify import file that adds one Shopify tag.

How to use:
1. Export products from Shopify/Matrixify as CSV or XLSX.
2. Run this script:
       python add_shopify_tag_rule.py
3. Answer the prompts for source tags, optional condition, target tag, input
   file, and output file.
4. Import the generated output file with Matrixify. The file contains only:
       Handle, Tags, Tags Command
   and every row uses Tags Command = MERGE.

The script never edits or overwrites your original export.
"""

from __future__ import annotations

import csv
import re
import sys
from collections import OrderedDict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterable


OUTPUT_COLUMNS = ["Handle", "Tags", "Tags Command"]
REPORT_EXAMPLE_LIMIT = 20


@dataclass
class ProductState:
    """Aggregated product-level state across one or more Matrixify rows."""

    handle: str
    matches_source_tag: bool = False
    matches_condition: bool = False
    already_has_target_tag: bool = False
    source_tags_seen: set[str] = field(default_factory=set)


def normalize_text(value: object) -> str:
    """Normalize text for case-insensitive comparisons with forgiving spaces."""
    if value is None:
        return ""
    return " ".join(str(value).strip().split()).casefold()


def display_text(value: object) -> str:
    """Convert spreadsheet values into clean display strings for output/reporting."""
    if value is None:
        return ""
    return str(value).strip()


def split_tags(tags_value: object) -> list[str]:
    """Split Shopify tags by comma and discard empty entries."""
    return [tag for tag in (normalize_text(part) for part in display_text(tags_value).split(",")) if tag]


def parse_comma_list(value: str) -> list[str]:
    """Parse comma-separated prompt input and normalize each entry."""
    items = [normalize_text(part) for part in value.split(",")]
    return [item for item in items if item]


def prompt_required(prompt: str) -> str:
    while True:
        value = clean_prompt_value(input(prompt))
        if value:
            return value
        print("Please enter a value.")


def clean_prompt_value(value: str) -> str:
    """Trim whitespace and matching quotes from pasted file paths or values."""
    value = value.strip()
    if len(value) >= 2 and value[0] == value[-1] and value[0] in {'"', "'"}:
        return value[1:-1].strip()
    return value


def prompt_yes_no(prompt: str) -> bool:
    while True:
        value = input(prompt).strip().casefold()
        if value in {"yes", "y"}:
            return True
        if value in {"no", "n"}:
            return False
        print("Please answer yes or no.")


def prompt_condition_mode() -> str:
    while True:
        value = input("Condition mode: equals/contains\n").strip().casefold()
        if value in {"equals", "contains"}:
            return value
        print("Please enter either equals or contains.")


def prompt_column(columns: list[str]) -> str:
    print("\nAvailable columns:")
    for index, column in enumerate(columns, start=1):
        print(f"{index}. {column}")

    while True:
        value = input("\nWhich column should the condition check?\n").strip()
        if value.isdigit():
            index = int(value)
            if 1 <= index <= len(columns):
                return columns[index - 1]

        for column in columns:
            if normalize_text(column) == normalize_text(value):
                return column

        print("Please enter a column name exactly as shown, or its number.")


def read_csv(path: Path) -> tuple[list[str], list[dict[str, object]]]:
    with path.open("r", encoding="utf-8-sig", newline="") as file:
        reader = csv.DictReader(file)
        columns = list(reader.fieldnames or [])
        rows = [dict(row) for row in reader]
    return columns, rows


def read_xlsx(path: Path) -> tuple[list[str], list[dict[str, object]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError(
            "XLSX support requires openpyxl. Install it with: pip install openpyxl"
        ) from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active

    rows_iter = worksheet.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if not header_row:
        return [], []

    columns = [display_text(cell) for cell in header_row]
    rows: list[dict[str, object]] = []
    for row in rows_iter:
        row_data = {column: row[index] if index < len(row) else None for index, column in enumerate(columns)}
        rows.append(row_data)

    return columns, rows


def read_export(path: Path) -> tuple[list[str], list[dict[str, object]]]:
    suffix = path.suffix.casefold()
    if suffix == ".csv":
        return read_csv(path)
    if suffix == ".xlsx":
        return read_xlsx(path)
    raise ValueError("Input file must be a .csv or .xlsx file.")


def write_csv(path: Path, rows: Iterable[dict[str, str]]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=OUTPUT_COLUMNS)
        writer.writeheader()
        writer.writerows(rows)


def write_xlsx(path: Path, rows: Iterable[dict[str, str]]) -> None:
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise RuntimeError(
            "XLSX output requires openpyxl. Install it with: pip install openpyxl"
        ) from exc

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Products"
    worksheet.append(OUTPUT_COLUMNS)
    for row in rows:
        worksheet.append([row[column] for column in OUTPUT_COLUMNS])
    workbook.save(path)


def write_import_file(path: Path, rows: list[dict[str, str]]) -> None:
    suffix = path.suffix.casefold()
    if suffix == ".csv":
        write_csv(path, rows)
        return
    if suffix == ".xlsx":
        write_xlsx(path, rows)
        return
    raise ValueError("Output file must end in .csv or .xlsx.")


def condition_matches(value: object, mode: str, expected_value: str) -> bool:
    actual = normalize_text(value)
    expected = normalize_text(expected_value)
    if mode == "equals":
        return actual == expected
    if mode == "contains":
        return expected in actual
    raise ValueError(f"Unsupported condition mode: {mode}")


def ensure_required_columns(columns: list[str]) -> None:
    missing = [column for column in ("Handle", "Tags") if column not in columns]
    if missing:
        raise ValueError(f"Missing required column(s): {', '.join(missing)}")


def aggregate_products(
    rows: list[dict[str, object]],
    source_tags: set[str],
    target_tag: str,
    condition_column: str | None,
    condition_mode: str | None,
    condition_value: str | None,
) -> OrderedDict[str, ProductState]:
    products: OrderedDict[str, ProductState] = OrderedDict()

    for row in rows:
        handle = display_text(row.get("Handle"))
        if not handle:
            continue

        product = products.setdefault(handle, ProductState(handle=handle))
        row_tags = set(split_tags(row.get("Tags")))
        matching_source_tags = row_tags.intersection(source_tags)

        if matching_source_tags:
            product.matches_source_tag = True
            product.source_tags_seen.update(matching_source_tags)

        if target_tag in row_tags:
            product.already_has_target_tag = True

        if condition_column is None:
            product.matches_condition = True
        elif condition_mode and condition_value is not None:
            if condition_matches(row.get(condition_column), condition_mode, condition_value):
                product.matches_condition = True

    return products


def default_report_path(output_path: Path) -> Path:
    return output_path.with_name(f"{output_path.stem}-review-report.txt")


def slugify(value: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "-", normalize_text(value)).strip("-")
    return slug or "tag"


def build_report(
    report_path: Path,
    input_path: Path,
    output_path: Path,
    total_rows_scanned: int,
    products: OrderedDict[str, ProductState],
    source_tags: set[str],
    target_tag_display: str,
    condition_column: str | None,
    condition_mode: str | None,
    condition_value: str | None,
    output_rows: list[dict[str, str]],
) -> None:
    source_matches = [product for product in products.values() if product.matches_source_tag]
    source_and_condition_matches = [
        product for product in source_matches if product.matches_condition
    ]
    skipped_existing_target = [
        product for product in source_and_condition_matches if product.already_has_target_tag
    ]
    written_handles = [row["Handle"] for row in output_rows]

    condition_summary = "None"
    if condition_column and condition_mode and condition_value is not None:
        condition_summary = f"{condition_column} {condition_mode} {condition_value}"

    lines = [
        "Shopify tag rule review report",
        "",
        f"Input file: {input_path}",
        f"Output file: {output_path}",
        f"Target tag: {target_tag_display}",
        f"Source tags: {', '.join(sorted(source_tags))}",
        f"Condition: {condition_summary}",
        "",
        f"total rows scanned: {total_rows_scanned}",
        f"number of products matching source tags: {len(source_matches)}",
        "number of products matching source tags + condition: "
        f"{len(source_and_condition_matches)}",
        "number of products skipped because they already had the target tag: "
        f"{len(skipped_existing_target)}",
        f"number of products written to output: {len(output_rows)}",
        "",
        "example matching handles:",
        *[f"- {handle}" for handle in written_handles[:REPORT_EXAMPLE_LIMIT]],
        "",
        "example skipped handles:",
        *[f"- {product.handle}" for product in skipped_existing_target[:REPORT_EXAMPLE_LIMIT]],
        "",
    ]

    report_path.write_text("\n".join(lines), encoding="utf-8")


def main() -> int:
    print("Shopify/Matrixify single-tag rule builder\n")

    input_file = prompt_required("Input file: ")
    input_path = Path(input_file).expanduser()
    if not input_path.exists():
        print(f"Input file not found: {input_path}", file=sys.stderr)
        return 1

    try:
        columns, rows = read_export(input_path)
        ensure_required_columns(columns)
    except Exception as exc:
        print(f"Could not read input file: {exc}", file=sys.stderr)
        return 1

    source_tag_input = prompt_required("\nEnter tags to look for, separated by commas:\n")
    source_tags = set(parse_comma_list(source_tag_input))
    if not source_tags:
        print("No valid source tags were entered.", file=sys.stderr)
        return 1

    condition_column = None
    condition_mode = None
    condition_value = None
    if prompt_yes_no("\nDo you want to add an additional condition? yes/no\n"):
        condition_column = prompt_column(columns)
        condition_mode = prompt_condition_mode()
        condition_value = prompt_required("What value should this column contain or equal?\n")

    target_tag_display = prompt_required("\nWhat tag should be added to matching products?\n")
    target_tag = normalize_text(target_tag_display)

    default_output = f"matrixify-add-{slugify(target_tag_display)}.csv"
    output_file = clean_prompt_value(input(f"\nOutput file [{default_output}]: ")) or default_output
    output_path = Path(output_file).expanduser()
    if output_path.exists():
        overwrite = prompt_yes_no(f"{output_path} already exists. Overwrite it? yes/no\n")
        if not overwrite:
            print("Stopped without writing output.")
            return 1

    products = aggregate_products(
        rows=rows,
        source_tags=source_tags,
        target_tag=target_tag,
        condition_column=condition_column,
        condition_mode=condition_mode,
        condition_value=condition_value,
    )

    output_rows = [
        {"Handle": product.handle, "Tags": target_tag_display, "Tags Command": "MERGE"}
        for product in products.values()
        if product.matches_source_tag
        and product.matches_condition
        and not product.already_has_target_tag
    ]

    try:
        write_import_file(output_path, output_rows)
        report_path = default_report_path(output_path)
        build_report(
            report_path=report_path,
            input_path=input_path,
            output_path=output_path,
            total_rows_scanned=len(rows),
            products=products,
            source_tags=source_tags,
            target_tag_display=target_tag_display,
            condition_column=condition_column,
            condition_mode=condition_mode,
            condition_value=condition_value,
            output_rows=output_rows,
        )
    except Exception as exc:
        print(f"Could not write output files: {exc}", file=sys.stderr)
        return 1

    print("\nDone.")
    print(f"Products written: {len(output_rows)}")
    print(f"Matrixify import file: {output_path}")
    print(f"Review report: {report_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
